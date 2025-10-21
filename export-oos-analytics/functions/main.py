"""Firebase Cloud Function for exporting out-of-stock analytics to Excel."""

from __future__ import annotations

import io
from datetime import datetime
from typing import Iterable, List

from firebase_functions import https_fn
from firebase_functions.options import set_global_options
from firebase_admin import initialize_app
from flask import jsonify, send_file
from google.cloud import bigquery
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

DATE_FORMAT = "%Y-%m-%d"
BIGQUERY_PROJECT = "incarts"
DATASET_TABLE = "incarts.analytics.out_of_stock_analytics"


# Limit concurrent instances for cost control (mirrors other exports)
set_global_options(max_instances=10)

# Initialize Firebase Admin SDK once per container
initialize_app()


@https_fn.on_request()
def export_out_of_stock_analytics(request: https_fn.Request) -> https_fn.Response:
    """HTTP function that exports out-of-stock analytics into an Excel workbook."""

    if request.method == "OPTIONS":
        cors_headers = {
            "Access-Control-Allow-Origin": "*",
            "Access-Control-Allow-Methods": "GET",
            "Access-Control-Allow-Headers": "Content-Type",
            "Access-Control-Max-Age": "3600",
        }
        response = https_fn.Response(status=204)
        response.headers.update(cors_headers)
        return response

    cors_headers = {"Access-Control-Allow-Origin": "*"}

    try:
        args = request.args
        start_date_str = args.get("start_date")
        end_date_str = args.get("end_date")
        project_id = args.get("project_id")
        link_name = args.get("link_name")
        slug = args.get("slug")

        missing_params = [p for p in ("start_date", "end_date", "project_id") if not args.get(p)]
        if missing_params:
            response = jsonify(
                {
                    "error": "Missing required parameters",
                    "missing": missing_params,
                    "example": (
                        "/export_out_of_stock_analytics?start_date=2025-05-01"
                        "&end_date=2025-05-31&project_id=your-project-id"
                    ),
                }
            )
            response.status_code = 400
            response.headers.update(cors_headers)
            return response

        try:
            start_date = datetime.strptime(start_date_str, DATE_FORMAT).date()
            end_date = datetime.strptime(end_date_str, DATE_FORMAT).date()
        except ValueError as exc:
            response = jsonify({"error": f"Invalid date format: {exc}"})
            response.status_code = 400
            response.headers.update(cors_headers)
            return response

        if start_date > end_date:
            response = jsonify({"error": "start_date cannot be after end_date"})
            response.status_code = 400
            response.headers.update(cors_headers)
            return response

        client = bigquery.Client(project=BIGQUERY_PROJECT)

        where_clause, query_params = build_filters(
            start_date, end_date, project_id, link_name, slug
        )

        summary = get_summary_metrics(client, where_clause, query_params)
        daily = get_out_of_stock_by_date(client, where_clause, query_params)
        by_state = get_out_of_stock_by_state(client, where_clause, query_params)
        substitutions = get_substitution_details(client, where_clause, query_params)

        workbook = Workbook()
        workbook.remove(workbook.active)

        create_summary_sheet(
            workbook,
            summary,
            start_date_str,
            end_date_str,
            project_id,
            link_name,
            slug,
        )
        create_daily_sheet(workbook, daily)
        create_state_sheet(workbook, by_state)
        create_substitution_sheet(workbook, substitutions)

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        filename = (
            f"out_of_stock_analytics_{project_id}_{start_date_str}_to_{end_date_str}.xlsx"
        )

        response = send_file(
            output,
            mimetype=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
            as_attachment=True,
            download_name=filename,
        )
        response.headers.update(cors_headers)
        return response

    except Exception as exc:  # pragma: no cover - surfaced via HTTP response
        response = jsonify({"error": str(exc)})
        response.status_code = 500
        response.headers.update(cors_headers)
        return response


def build_filters(
    start_date,
    end_date,
    project_id,
    link_name: str | None,
    slug: str | None,
):
    """Construct WHERE clause and BigQuery parameters for shared filtering."""

    where_conditions: List[str] = [
        "date BETWEEN @start_date AND @end_date",
        "project_id = @project_id",
    ]

    params: List[bigquery.ScalarQueryParameter] = [
        bigquery.ScalarQueryParameter("start_date", "DATE", start_date),
        bigquery.ScalarQueryParameter("end_date", "DATE", end_date),
        bigquery.ScalarQueryParameter("project_id", "STRING", project_id),
    ]

    if link_name:
        where_conditions.append("link_name = @link_name")
        params.append(bigquery.ScalarQueryParameter("link_name", "STRING", link_name))

    if slug:
        where_conditions.append("short_id = @slug")
        params.append(bigquery.ScalarQueryParameter("slug", "STRING", slug))

    return " AND ".join(where_conditions), params


def get_summary_metrics(client, where_clause: str, params: List) -> bigquery.table.Row:
    """Return out-of-stock counts and distinct geography impact."""

    query = f"""
    SELECT
      COUNT(*) AS out_of_stock_count,
      COUNT(DISTINCT state) AS states_affected,
      COUNT(DISTINCT zip_code) AS zip_codes_affected,
      ANY_VALUE(project_name) AS project_name
    FROM `{DATASET_TABLE}`
    WHERE {where_clause}
    """

    job = client.query(query, job_config=bigquery.QueryJobConfig(query_parameters=params))
    return list(job.result())[0]


def get_out_of_stock_by_date(
    client, where_clause: str, params: List
) -> Iterable[bigquery.table.Row]:
    """Return daily out-of-stock counts."""

    query = f"""
    SELECT
      DATE(date) AS date,
      COUNT(*) AS count
    FROM `{DATASET_TABLE}`
    WHERE {where_clause}
    GROUP BY date
    ORDER BY date ASC
    """

    job = client.query(query, job_config=bigquery.QueryJobConfig(query_parameters=params))
    return list(job.result())


def get_out_of_stock_by_state(
    client, where_clause: str, params: List
) -> Iterable[bigquery.table.Row]:
    """Return counts of out-of-stock events by geography."""

    query = f"""
    SELECT
      state,
      city,
      COUNT(*) AS count
    FROM `{DATASET_TABLE}`
    WHERE {where_clause}
    GROUP BY state, city
    ORDER BY count DESC, state ASC, city ASC
    """

    job = client.query(query, job_config=bigquery.QueryJobConfig(query_parameters=params))
    return list(job.result())


def get_substitution_details(
    client, where_clause: str, params: List
) -> Iterable[bigquery.table.Row]:
    """Return substitution breakdown limited to top 25 rows."""

    query = f"""
    SELECT
      DATE(date) AS date,
      primary_product_name,
      replacement_product_name,
      substitution_reason,
      COUNT(*) AS count
    FROM `{DATASET_TABLE}`
    WHERE {where_clause}
    GROUP BY date, primary_product_name, replacement_product_name, substitution_reason
    ORDER BY count DESC, date ASC,
             primary_product_name ASC, replacement_product_name ASC,
             substitution_reason ASC
    LIMIT 25
    """

    job = client.query(query, job_config=bigquery.QueryJobConfig(query_parameters=params))
    return list(job.result())


def create_summary_sheet(
    workbook: Workbook,
    data,
    start_date: str,
    end_date: str,
    project_id: str,
    link_name: str | None,
    slug: str | None,
) -> None:
    """Create a summary sheet containing KPI cards."""

    ws = workbook.create_sheet("Summary")

    ws["A1"] = "Out of Stock Analytics Summary"
    ws["A1"].font = Font(size=16, bold=True, color="1F4E78")

    ws["A2"] = f"Project: {data.project_name or project_id}"
    ws["A2"].font = Font(size=11, italic=True)

    ws["A3"] = f"Period: {start_date} to {end_date}"
    ws["A3"].font = Font(size=11, italic=True)

    info_row = 4
    if link_name:
        ws[f"A{info_row}"] = f"Link Name: {link_name}"
        ws[f"A{info_row}"].font = Font(size=11, italic=True)
        info_row += 1
    if slug:
        ws[f"A{info_row}"] = f"Slug: {slug}"
        ws[f"A{info_row}"].font = Font(size=11, italic=True)
        info_row += 1

    header_row = info_row + 1
    headers = ["Metric", "Value"]
    for col_index, header in enumerate(headers, start=1):
        cell = ws.cell(header_row, col_index, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4472C4")
        cell.alignment = Alignment(horizontal="center")

    metrics_start_row = header_row + 1
    metrics = [
        ("Out of Stock Count", data.out_of_stock_count),
        ("States Affected", data.states_affected),
        ("Zip Codes Affected", data.zip_codes_affected),
    ]

    for row_offset, (metric, value) in enumerate(metrics):
        row_idx = metrics_start_row + row_offset
        ws.cell(row_idx, 1, metric)
        ws.cell(row_idx, 2, value)
        if row_offset % 2 == 1:
            for col_idx in range(1, 3):
                ws.cell(row_idx, col_idx).fill = PatternFill(
                    "solid", fgColor="D9E1F2"
                )

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 20


def create_daily_sheet(workbook: Workbook, rows: Iterable) -> None:
    """Add a sheet with out-of-stock counts by date."""

    ws = workbook.create_sheet("Out of Stock by Date")
    headers = ["Date", "Count"]
    _write_table(ws, headers, rows, row_formatter=_format_date_row)


def create_state_sheet(workbook: Workbook, rows: Iterable) -> None:
    """Add a sheet with geographic breakdown."""

    ws = workbook.create_sheet("Out of Stock by State")
    headers = ["State", "City", "Count"]
    _write_table(ws, headers, rows)


def create_substitution_sheet(workbook: Workbook, rows: Iterable) -> None:
    """Add a sheet showing top substitution combinations."""

    ws = workbook.create_sheet("Substitution Details")
    headers = [
        "Date",
        "Primary Product",
        "Replacement Product",
        "Substitution Reason",
        "Count",
    ]
    _write_table(ws, headers, rows, row_formatter=_format_substitution_row)


def _write_table(ws, headers: List[str], rows: Iterable, row_formatter=None) -> None:
    """Utility to render headers and rows onto a worksheet."""

    for col_index, header in enumerate(headers, start=1):
        cell = ws.cell(1, col_index, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4472C4")
        cell.alignment = Alignment(horizontal="center")

    for row_index, row in enumerate(rows, start=2):
        values = row_formatter(row) if row_formatter else tuple(row)
        for col_index, value in enumerate(values, start=1):
            ws.cell(row_index, col_index, value)

    for col_index in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_index)].width = 25


def _format_date_row(row) -> tuple:
    """Format rows containing a date and count."""

    return (
        row.date.strftime(DATE_FORMAT) if getattr(row, "date", None) else None,
        row.count,
    )


def _format_substitution_row(row) -> tuple:
    """Format substitution detail records for Excel output."""

    return (
        row.date.strftime(DATE_FORMAT) if getattr(row, "date", None) else None,
        row.primary_product_name,
        row.replacement_product_name,
        row.substitution_reason,
        row.count,
    )
