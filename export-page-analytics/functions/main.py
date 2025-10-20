"""
@fileoverview Firebase Cloud Function for exporting page analytics data to Excel spreadsheets
@description This module provides HTTP endpoint for generating comprehensive analytics reports
             with multiple sheets covering summary metrics, daily breakdowns, page performance,
             traffic sources, geographic distribution, device analytics, and click-through details.
             Data is fetched from BigQuery and formatted into a styled Excel workbook.
@module export-page-analytics
@example
    # Request analytics export for a date range
    GET /export_page_analytics?start_date=2025-01-01&end_date=2025-12-31&project_id=my-project
@related BigQuery tables: dashboard_summary_with_traffic_pages_v4, pages_geographic_analytics_v4,
         pages_device_analytics_v4, dashboard_click_through_details_pages_v4
"""

import io

from firebase_functions import https_fn
from google.cloud import bigquery
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from flask import jsonify, send_file

@https_fn.on_request()
def export_page_analytics(request: https_fn.Request) -> https_fn.Response:
    """
    Export page analytics to Excel spreadsheet
    
    Query Parameters:
        start_date (str): Start date in YYYY-MM-DD format (required)
        end_date (str): End date in YYYY-MM-DD format (required)
        project_id (str): Filter by project ID (optional)
        page_slug (str): Filter by page slug (optional)
        source (str): Filter by traffic source (optional)
        medium (str): Filter by traffic medium (optional)
    
    Returns:
        Excel file download with multiple sheets of analytics data
    """
    
    # Set CORS headers
    if request.method == 'OPTIONS':
        cors_headers = {
            'Access-Control-Allow-Origin': '*',
            'Access-Control-Allow-Methods': 'GET',
            'Access-Control-Allow-Headers': 'Content-Type',
            'Access-Control-Max-Age': '3600',
        }
        response = https_fn.Response(status=204)
        response.headers.update(cors_headers)
        return response

    cors_headers = {'Access-Control-Allow-Origin': '*'}
    
    try:
        # Parse query parameters
        args = request.args
        start_date = args.get('start_date')
        end_date = args.get('end_date')
        project_id = args.get('project_id')
        page_slug = args.get('page_slug')
        source = args.get('source')
        medium = args.get('medium')
        
        # Validate required parameters
        if not start_date or not end_date:
            response = jsonify({
                'error': 'start_date and end_date are required',
                'example': '/export_page_analytics?start_date=2025-01-01&end_date=2025-12-31'
            })
            response.status_code = 400
            response.headers.update(cors_headers)
            return response
        
        # Initialize BigQuery client
        client = bigquery.Client(project='incarts')
        
        # Build WHERE clause
        where_conditions = [f"date BETWEEN '{start_date}' AND '{end_date}'"]
        if project_id:
            where_conditions.append(f"project_id = '{project_id}'")
        if page_slug:
            where_conditions.append(f"page_slug = '{page_slug}'")
        if source:
            where_conditions.append(f"source = '{source}'")
        if medium:
            where_conditions.append(f"medium = '{medium}'")
        
        where_clause = " AND ".join(where_conditions)
        
        # Create Excel workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Sheet 1: Summary Metrics
        summary_data = get_summary_metrics(client, where_clause)
        create_summary_sheet(wb, summary_data, start_date, end_date)
        
        # Sheet 2: Daily Breakdown
        daily_data = get_daily_breakdown(client, where_clause)
        create_daily_sheet(wb, daily_data)
        
        # Sheet 3: Page Performance
        page_data = get_page_performance(client, where_clause)
        create_page_sheet(wb, page_data)
        
        # Sheet 4: Traffic Sources
        traffic_data = get_traffic_sources(client, where_clause)
        create_traffic_sheet(wb, traffic_data)
        
        # Sheet 5: Geographic Distribution
        geo_data = get_geographic_data(client, where_clause)
        create_geographic_sheet(wb, geo_data)
        
        # Sheet 6: Device Breakdown
        device_data = get_device_breakdown(client, where_clause)
        create_device_sheet(wb, device_data)
        
        # Sheet 7: Click Details (if any)
        click_data = get_click_details(client, where_clause)
        if click_data:
            create_clicks_sheet(wb, click_data)
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generate filename
        filename = f"page_analytics_{start_date}_to_{end_date}.xlsx"
        
        # Return file
        response = send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        response.headers.update(cors_headers)
        return response

    except Exception as e:
        response = jsonify({'error': str(e)})
        response.status_code = 500
        response.headers.update(cors_headers)
        return response


def get_summary_metrics(client, where_clause):
    """Get overall summary metrics"""
    query = f"""
    SELECT 
      COUNT(DISTINCT page_slug) as total_pages,
      COUNT(DISTINCT project_id) as total_projects,
      SUM(total_page_views) as total_page_views,
      SUM(total_users) as total_users,
      SUM(total_sessions) as total_sessions,
      SUM(total_clicks) as total_clicks,
      ROUND(AVG(avg_session_duration_seconds), 2) as avg_session_duration,
      ROUND(AVG(bounce_rate_pct), 2) as avg_bounce_rate,
      ROUND(SAFE_DIVIDE(SUM(total_clicks), SUM(total_page_views)) * 100, 2) as overall_ctr
    FROM `incarts.analytics.dashboard_summary_with_traffic_pages_v4`
    WHERE {where_clause}
    """
    return list(client.query(query).result())[0]


def get_daily_breakdown(client, where_clause):
    """Get daily breakdown of metrics"""
    query = f"""
    SELECT 
      date,
      SUM(total_page_views) as page_views,
      SUM(total_users) as users,
      SUM(total_sessions) as sessions,
      SUM(total_clicks) as clicks,
      ROUND(SAFE_DIVIDE(SUM(total_clicks), SUM(total_page_views)) * 100, 2) as ctr_pct,
      COUNT(DISTINCT page_slug) as active_pages
    FROM `incarts.analytics.dashboard_summary_with_traffic_pages_v4`
    WHERE {where_clause}
    GROUP BY date
    ORDER BY date ASC
    """
    return list(client.query(query).result())


def get_page_performance(client, where_clause):
    """Get page-level performance metrics"""
    query = f"""
    SELECT 
      page_slug,
      project_name,
      SUM(total_page_views) as total_views,
      SUM(total_users) as total_users,
      SUM(total_sessions) as total_sessions,
      SUM(total_clicks) as total_clicks,
      ROUND(SAFE_DIVIDE(SUM(total_clicks), SUM(total_page_views)) * 100, 2) as ctr_pct,
      ROUND(AVG(avg_session_duration_seconds), 2) as avg_session_duration,
      ROUND(AVG(bounce_rate_pct), 2) as avg_bounce_rate,
      COUNT(DISTINCT date) as days_active
    FROM `incarts.analytics.dashboard_summary_with_traffic_pages_v4`
    WHERE {where_clause}
    GROUP BY page_slug, project_name
    ORDER BY total_views DESC
    """
    return list(client.query(query).result())


def get_traffic_sources(client, where_clause):
    """Get traffic source breakdown"""
    query = f"""
    SELECT 
      source,
      medium,
      SUM(total_page_views) as page_views,
      SUM(total_users) as users,
      SUM(total_sessions) as sessions,
      COUNT(DISTINCT page_slug) as unique_pages
    FROM `incarts.analytics.dashboard_summary_with_traffic_pages_v4`
    WHERE {where_clause}
    GROUP BY source, medium
    ORDER BY page_views DESC
    """
    return list(client.query(query).result())


def get_geographic_data(client, where_clause):
    """Get geographic distribution"""
    query = f"""
    SELECT 
      country,
      state,
      city,
      SUM(page_views) as total_page_views,
      SUM(users) as total_users,
      COUNT(DISTINCT page_slug) as unique_pages
    FROM `incarts.analytics.pages_geographic_analytics_v4`
    WHERE {where_clause}
      AND city != '(not set)'
    GROUP BY country, state, city
    ORDER BY total_page_views DESC
    LIMIT 100
    """
    return list(client.query(query).result())


def get_device_breakdown(client, where_clause):
    """Get device category breakdown"""
    query = f"""
    SELECT 
      device_category,
      SUM(page_views) as total_page_views,
      SUM(users) as total_users,
      SUM(sessions) as total_sessions,
      ROUND(AVG(avg_session_duration_seconds), 2) as avg_session_duration,
      ROUND(AVG(bounce_rate_pct), 2) as avg_bounce_rate
    FROM `incarts.analytics.pages_device_analytics_v4`
    WHERE {where_clause}
    GROUP BY device_category
    ORDER BY total_page_views DESC
    """
    return list(client.query(query).result())


def get_click_details(client, where_clause):
    """Get click through details if available"""
    query = f"""
    SELECT 
      page_slug,
      destination_url,
      event_name,
      SUM(total_clicks) as clicks,
      MAX(page_views_that_day) as page_views,
      ROUND(AVG(link_ctr_pct), 2) as avg_ctr
    FROM `incarts.analytics.dashboard_click_through_details_pages_v4`
    WHERE {where_clause}
    GROUP BY page_slug, destination_url, event_name
    ORDER BY clicks DESC
    LIMIT 500
    """
    results = list(client.query(query).result())
    return results if results else None


def create_summary_sheet(wb, data, start_date, end_date):
    """Create summary metrics sheet"""
    ws = wb.create_sheet('Summary')
    
    # Title
    ws['A1'] = 'Page Analytics Summary Report'
    ws['A1'].font = Font(size=16, bold=True, color='1F4E78')
    ws['A2'] = f'Period: {start_date} to {end_date}'
    ws['A2'].font = Font(size=11, italic=True)
    
    # Headers
    headers = ['Metric', 'Value']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(4, col, header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
    
    # Data
    metrics = [
        ('Total Pages', data.total_pages),
        ('Total Projects', data.total_projects),
        ('Total Page Views', data.total_page_views),
        ('Total Users', data.total_users),
        ('Total Sessions', data.total_sessions),
        ('Total Clicks', data.total_clicks),
        ('Average Session Duration (seconds)', data.avg_session_duration),
        ('Average Bounce Rate (%)', data.avg_bounce_rate),
        ('Overall CTR (%)', data.overall_ctr),
    ]
    
    for idx, (metric, value) in enumerate(metrics, 5):
        ws.cell(idx, 1, metric)
        ws.cell(idx, 2, value)
        if idx % 2 == 0:
            ws.cell(idx, 1).fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
            ws.cell(idx, 2).fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    
    # Column widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 20


def create_daily_sheet(wb, data):
    """Create daily breakdown sheet"""
    ws = wb.create_sheet('Daily Breakdown')
    
    # Headers
    headers = ['Date', 'Page Views', 'Users', 'Sessions', 'Clicks', 'CTR (%)', 'Active Pages']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
    
    # Data
    for row_idx, row in enumerate(data, 2):
        ws.cell(row_idx, 1, row.date.strftime('%Y-%m-%d') if row.date else '')
        ws.cell(row_idx, 2, row.page_views)
        ws.cell(row_idx, 3, row.users)
        ws.cell(row_idx, 4, row.sessions)
        ws.cell(row_idx, 5, row.clicks)
        ws.cell(row_idx, 6, row.ctr_pct)
        ws.cell(row_idx, 7, row.active_pages)
    
    # Format and auto-width
    for col in range(1, 8):
        ws.column_dimensions[get_column_letter(col)].width = 15


def create_page_sheet(wb, data):
    """Create page performance sheet"""
    ws = wb.create_sheet('Page Performance')
    
    # Headers
    headers = ['Page Slug', 'Project', 'Views', 'Users', 'Sessions', 'Clicks', 
               'CTR (%)', 'Avg Duration (s)', 'Bounce Rate (%)', 'Days Active']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
    
    # Data
    for row_idx, row in enumerate(data, 2):
        ws.cell(row_idx, 1, row.page_slug)
        ws.cell(row_idx, 2, row.project_name)
        ws.cell(row_idx, 3, row.total_views)
        ws.cell(row_idx, 4, row.total_users)
        ws.cell(row_idx, 5, row.total_sessions)
        ws.cell(row_idx, 6, row.total_clicks)
        ws.cell(row_idx, 7, row.ctr_pct)
        ws.cell(row_idx, 8, row.avg_session_duration)
        ws.cell(row_idx, 9, row.avg_bounce_rate)
        ws.cell(row_idx, 10, row.days_active)
    
    # Auto-width
    for col in range(1, 11):
        ws.column_dimensions[get_column_letter(col)].width = 18


def create_traffic_sheet(wb, data):
    """Create traffic sources sheet"""
    ws = wb.create_sheet('Traffic Sources')
    
    # Headers
    headers = ['Source', 'Medium', 'Page Views', 'Users', 'Sessions', 'Unique Pages']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
    
    # Data
    for row_idx, row in enumerate(data, 2):
        ws.cell(row_idx, 1, row.source)
        ws.cell(row_idx, 2, row.medium)
        ws.cell(row_idx, 3, row.page_views)
        ws.cell(row_idx, 4, row.users)
        ws.cell(row_idx, 5, row.sessions)
        ws.cell(row_idx, 6, row.unique_pages)
    
    # Auto-width
    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 20


def create_geographic_sheet(wb, data):
    """Create geographic distribution sheet"""
    ws = wb.create_sheet('Geographic Distribution')
    
    # Headers
    headers = ['Country', 'State', 'City', 'Page Views', 'Users', 'Unique Pages']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
    
    # Data
    for row_idx, row in enumerate(data, 2):
        ws.cell(row_idx, 1, row.country)
        ws.cell(row_idx, 2, row.state if row.state else '')
        ws.cell(row_idx, 3, row.city)
        ws.cell(row_idx, 4, row.total_page_views)
        ws.cell(row_idx, 5, row.total_users)
        ws.cell(row_idx, 6, row.unique_pages)
    
    # Auto-width
    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 20


def create_device_sheet(wb, data):
    """Create device breakdown sheet"""
    ws = wb.create_sheet('Device Breakdown')
    
    # Headers
    headers = ['Device Category', 'Page Views', 'Users', 'Sessions', 
               'Avg Duration (s)', 'Bounce Rate (%)']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
    
    # Data
    for row_idx, row in enumerate(data, 2):
        ws.cell(row_idx, 1, row.device_category)
        ws.cell(row_idx, 2, row.total_page_views)
        ws.cell(row_idx, 3, row.total_users)
        ws.cell(row_idx, 4, row.total_sessions)
        ws.cell(row_idx, 5, row.avg_session_duration)
        ws.cell(row_idx, 6, row.avg_bounce_rate)
    
    # Auto-width
    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 20


def create_clicks_sheet(wb, data):
    """Create click details sheet"""
    ws = wb.create_sheet('Click Details')
    
    # Headers
    headers = ['Page Slug', 'Destination URL', 'Event Name', 'Total Clicks', 
               'Page Views', 'CTR (%)']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
    
    # Data
    for row_idx, row in enumerate(data, 2):
        ws.cell(row_idx, 1, row.page_slug)
        ws.cell(row_idx, 2, row.destination_url)
        ws.cell(row_idx, 3, row.event_name)
        ws.cell(row_idx, 4, row.clicks)
        ws.cell(row_idx, 5, row.page_views)
        ws.cell(row_idx, 6, row.avg_ctr)
    
    # Auto-width
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 50
    for col in range(3, 7):
        ws.column_dimensions[get_column_letter(col)].width = 15
